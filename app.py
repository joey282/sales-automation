import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
from thefuzz import process, fuzz
from streamlit_gsheets import GSheetsConnection
import io
import os

# --- 1. ตั้งค่าหน้าจอและระบบความปลอดภัย ---
st.set_page_config(page_title="Sales Summary", layout="wide")

def check_password():
    if "password_correct" not in st.session_state:
        st.session_state.password_correct = False
    if not st.session_state.password_correct:
        st.title("🔐 กรุณาใส่รหัสผ่านเพื่อใช้งาน")
        password = st.text_input("Password", type="password")
        if st.button("Login"):
            if password == "1234": # <--- รหัสผ่านเข้าหน้าเว็บ
                st.session_state.password_correct = True
                st.rerun()
            else:
                st.error("รหัสผ่านไม่ถูกต้อง")
        return False
    return True

if not check_password():
    st.stop()

# --- 2. เชื่อมต่อ Google Sheets (ฐานข้อมูลหลัก) ---
# ระบุ URL ของ Google Sheets ของคุณที่นี่
SHEET_URL = "https://docs.google.com/spreadsheets/d/1cM7lKy8jq3wcjvBz3fH1tAmR9jjKZ2QDwETPDeDdlME/edit?usp=sharing"

try:
    conn = st.connection("gsheets", type=GSheetsConnection)
    # ใช้ ttl=0 เพื่อให้ดึงข้อมูลใหม่ทุกครั้งที่ Refresh
    df_mapping = conn.read(spreadsheet=SHEET_URL, ttl=0)
    keywords = df_mapping.iloc[:, 0].dropna().tolist()
    st.sidebar.success("✅ เชื่อมต่อฐานข้อมูลเมนูแล้ว")
except Exception as e:
    st.sidebar.error("❌ ไม่สามารถเชื่อมต่อ Google Sheets ได้")
    st.stop()

# --- 3. ฟังก์ชันดึงข้อมูลจาก HTML ---
def extract_html_data(uploaded_file):
    content = uploaded_file.read().decode("utf-8")
    soup = BeautifulSoup(content, 'html.parser')
    rows_data = []
    file_display_name = os.path.splitext(uploaded_file.name)[0]
    
    for row in soup.find_all('tr'):
        cols = [ele.text.strip() for ele in row.find_all(['td', 'th'])]
        if len(cols) >= 5:
            menu_name = cols[2].replace('\n', ' ').strip()
            qty_raw = cols[3].strip()
            if qty_raw.isdigit():
                rows_data.append({
                    'Menu Name': menu_name,
                    'Quantity': int(qty_raw),
                    'Source File': file_display_name
                })
    return pd.DataFrame(rows_data)

# --- 4. ส่วนแสดงผลหลัก ---
st.title("📊 ระบบ สรุปยอดขายรายไฟล์ (Version 5.0)")
st.write(f"ฐานข้อมูลเมนูปัจจุบันมี: {len(keywords)} รายการ")

if st.sidebar.button("🔄 รีเฟรชข้อมูลจาก Google Sheets"):
    st.cache_data.clear()
    st.rerun()

uploaded_files = st.sidebar.file_uploader("📥 อัปโหลดไฟล์ HTML", type=['html'], accept_multiple_files=True)

if uploaded_files:
    all_processed_data = []
    all_unmatched = []

    # ประมวลผลแต่ละไฟล์
    for uploaded_file in uploaded_files:
        df_sales = extract_html_data(uploaded_file)
        
        if not df_sales.empty:
            matched_categories = []
            for menu in df_sales['Menu Name']:
                match, score = process.extractOne(menu, keywords, scorer=fuzz.token_sort_ratio)
                if score >= 60:
                    cat = df_mapping[df_mapping.iloc[:, 0] == match].iloc[0, 1]
                    matched_categories.append(cat)
                else:
                    matched_categories.append("Unknown")
                    if menu not in all_unmatched: all_unmatched.append(menu)
            
            df_sales['Category'] = matched_categories
            all_processed_data.append(df_sales)

            # --- แสดงข้อมูลดิบรายไฟล์ใน Expander ---
            file_name = os.path.splitext(uploaded_file.name)[0]
            with st.expander(f"📁 ตารางข้อมูลดิบและสรุปประเภท: {file_name}"):
                # ตารางสรุปประเภทของไฟล์นี้
                st.subheader("สรุปแยกตามประเภท")
                file_summary = df_sales.groupby('Category')['Quantity'].sum().reset_index()

                # เพิ่มแถว Total ท้ายตารางสรุปรายไฟล์
                total_qty = file_summary['Quantity'].sum()
                sum_total_row = pd.DataFrame([['ยอดรวมทั้งหมด (Total)', total_qty]], columns=['Category', 'Quantity'])
                file_summary_with_total = pd.concat([file_summary, sum_total_row], ignore_index=True)
                st.table(file_summary_with_total)

                # --- ตารางเช็ทเมนู (เมนูที่มี " + " ในชื่อ) ---
                st.subheader("🍱 รายการเช็ทเมนู")
                df_set_menu = df_sales[df_sales['Menu Name'].str.contains(r'\+', na=False)]
                if not df_set_menu.empty:
                    set_menu_summary = df_set_menu.groupby('Menu Name')['Quantity'].sum().reset_index()
                    set_menu_summary = set_menu_summary.sort_values('Quantity', ascending=False).reset_index(drop=True)
                    set_total = set_menu_summary['Quantity'].sum()
                    set_total_row = pd.DataFrame([['ยอดรวมเช็ทเมนู (Total)', set_total]], columns=['Menu Name', 'Quantity'])
                    set_menu_display = pd.concat([set_menu_summary, set_total_row], ignore_index=True)
                    st.dataframe(set_menu_display, use_container_width=True)
                else:
                    st.info("ไม่พบเช็ทเมนูในไฟล์นี้")

                # ตารางข้อมูลดิบ (ชื่อเมนู + ประเภท)
                st.subheader("รายการขายทั้งหมด")
                # เพิ่มแถว Total ท้ายตารางข้อมูลดิบ
                raw_total_row = pd.DataFrame([['รวม', total_qty, '', '']], columns=['Menu Name', 'Quantity', 'Source File', 'Category'])
                df_display = pd.concat([df_sales, raw_total_row], ignore_index=True).fillna('')
                st.dataframe(df_display, use_container_width=True)

    # --- 5. ตารางรวมเปรียบเทียบ (Pivot Table) ---
    if all_processed_data:
        final_df = pd.concat(all_processed_data, ignore_index=True)
        
        pivot_table = final_df.pivot_table(
            index='Category',
            columns='Source File',
            values='Quantity',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        # เพิ่มยอดรวมแนวนอน (Total Per Category)
        pivot_table['ยอดรวมทุกไฟล์'] = pivot_table.iloc[:, 1:].sum(axis=1)

        # --- เพิ่มยอดรวมแนวตั้ง (Total Row) ท้ายตาราง Pivot ---
        total_values = ['ยอดรวมทั้งหมด (Total)']
        for col in pivot_table.columns[1:]:
            total_values.append(pivot_table[col].sum())
        
        pivot_total_row = pd.DataFrame([total_values], columns=pivot_table.columns)
        pivot_table_final = pd.concat([pivot_table, pivot_total_row], ignore_index=True)

        st.divider()
        st.header("📊 ตารางสรุปเปรียบเทียบทุกไฟล์")
        st.dataframe(pivot_table_final, use_container_width=True)

        # --- เตรียมข้อมูลแยกตามสาขา (สำหรับชีท Set Menu และ Top 10) ---
        branches = list(final_df['Source File'].unique())

        set_menu_per_branch = {}
        top10_per_branch = {}
        for branch in branches:
            df_branch = final_df[final_df['Source File'] == branch]

            df_set = df_branch[df_branch['Menu Name'].str.contains(r'\+', na=False)]
            set_summary = df_set.groupby('Menu Name')['Quantity'].sum().reset_index()
            set_summary = set_summary.sort_values('Quantity', ascending=False).reset_index(drop=True)
            set_summary.insert(0, 'ອັນດັບ', range(1, len(set_summary) + 1))
            set_menu_per_branch[branch] = set_summary

            df_branch_filtered = df_branch[
                df_branch['Category'].astype(str).str.strip().str.lower().isin(['iced', 'hot', 'smt'])
            ]
            top10 = df_branch_filtered.groupby('Menu Name')['Quantity'].sum().reset_index()
            top10 = top10.sort_values('Quantity', ascending=False).head(10).reset_index(drop=True)
            top10.insert(0, 'ອັນດັບ', range(1, len(top10) + 1))
            top10_per_branch[branch] = top10

        def write_branch_horizontal(ws, branch_data_dict, branches_list):
            """เขียนข้อมูลแบบเรียงข้างกัน: แต่ละสาขาใช้ 3 คอลัมน์ (ອັນດັບ, Menu Name, Quantity)"""
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            title_font = Font(bold=True, size=12)
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center")
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for idx, branch in enumerate(branches_list):
                start_col = idx * 3 + 1
                df = branch_data_dict[branch]

                # Row 1: ชื่อสาขา (merge 3 คอลัมน์)
                ws.cell(row=1, column=start_col, value=branch)
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 2)
                title_cell = ws.cell(row=1, column=start_col)
                title_cell.font = title_font
                title_cell.alignment = center

                # Row 2: หัวคอลัมน์
                headers = ['ອັນດັບ', 'Menu Name', 'Quantity']
                for j, h in enumerate(headers):
                    c = ws.cell(row=2, column=start_col + j, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = center
                    c.border = border

                # Rows 3+: ข้อมูล
                for r_idx, row in df.iterrows():
                    for j, key in enumerate(['ອັນດັບ', 'Menu Name', 'Quantity']):
                        c = ws.cell(row=3 + r_idx, column=start_col + j, value=row[key])
                        c.border = border
                        if key in ('ອັນດັບ', 'Quantity'):
                            c.alignment = center

                # ปรับความกว้างคอลัมน์
                ws.column_dimensions[get_column_letter(start_col)].width = 8
                ws.column_dimensions[get_column_letter(start_col + 1)].width = 32
                ws.column_dimensions[get_column_letter(start_col + 2)].width = 12

        # ปุ่มดาวน์โหลด Excel
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            pivot_table_final.to_excel(writer, index=False, sheet_name='Summary_Report')

            wb = writer.book

            # ชีท Set_Menu_By_Branch
            ws_set = wb.create_sheet('Set_Menu_By_Branch')
            if any(not df.empty for df in set_menu_per_branch.values()):
                write_branch_horizontal(ws_set, set_menu_per_branch, branches)
            else:
                ws_set.cell(row=1, column=1, value='ไม่พบเช็ทเมนู')

            # ชีท Top10_By_Branch
            ws_top = wb.create_sheet('Top10_By_Branch')
            write_branch_horizontal(ws_top, top10_per_branch, branches)
        
        st.download_button(
            label="📥 ดาวน์โหลดไฟล์สรุปรายงาน (Excel)",
            data=buffer.getvalue(),
            file_name="Sales_Master_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if all_unmatched:
            st.warning("⚠️ พบเมนูใหม่ที่ไม่มีในฐานข้อมูล (กรุณาเพิ่มใน Google Sheets)")
            with st.expander("คลิกเพื่อดูเมนูที่ต้องเพิ่ม"):
                for item in all_unmatched:
                    st.write(f"- {item}")
else:
    st.info("💡 เริ่มต้นโดยการอัปโหลดไฟล์ HTML ที่แถบด้านซ้าย")
